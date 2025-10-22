"""
Views for Worksheet Generator functionality.
Handles equipment types, worksheet templates, and worksheet generation.
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.db.models import Q
from django.conf import settings
import io
import os
from datetime import datetime

# Import XLSX library for Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from .models import EquipmentType, WorksheetTemplate, CustomerMembership
from .serializers import EquipmentTypeSerializer, WorksheetTemplateSerializer


class EquipmentTypeViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing equipment types.
    Provides CRUD operations for equipment types.
    """
    serializer_class = EquipmentTypeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Get all active equipment types, optionally filtered by category"""
        queryset = EquipmentType.objects.filter(is_active=True)

        # Optional filtering by category
        category = self.request.query_params.get('category', None)
        if category:
            queryset = queryset.filter(category=category)

        return queryset.order_by('category', 'display_order', 'name')

    @action(detail=False, methods=['get'])
    def categories(self, request):
        """Get list of all equipment categories"""
        categories = EquipmentType.CATEGORY_CHOICES
        return Response([
            {'value': cat[0], 'label': cat[1]}
            for cat in categories
        ])

    @action(detail=False, methods=['get'])
    def by_category(self, request):
        """Get equipment types grouped by category"""
        categories = {}
        equipment_types = self.get_queryset()

        for eq_type in equipment_types:
            category = eq_type.category
            if category not in categories:
                categories[category] = []

            serializer = self.get_serializer(eq_type)
            categories[category].append(serializer.data)

        return Response(categories)


class WorksheetTemplateViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing worksheet templates.
    Provides CRUD operations for worksheet templates.
    """
    serializer_class = WorksheetTemplateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Get templates accessible to the user"""
        user = self.request.user

        # Get customer IDs the user has access to
        customer_ids = CustomerMembership.objects.filter(
            user=user
        ).values_list('customer_id', flat=True)

        # Get templates: global templates + user's templates + customer templates
        queryset = WorksheetTemplate.objects.filter(
            Q(is_global=True) |
            Q(user=user) |
            Q(customer_id__in=customer_ids)
        )

        # Optional filtering
        customer_id = self.request.query_params.get('customer', None)
        if customer_id:
            queryset = queryset.filter(
                Q(is_global=True) | Q(customer_id=customer_id)
            )

        return queryset.select_related('customer', 'user').prefetch_related('equipment_types')

    def perform_create(self, serializer):
        """Create template with current user"""
        serializer.save(user=self.request.user)

    @action(detail=False, methods=['post'])
    def generate_worksheet(self, request):
        """
        Generate an Excel worksheet based on the provided configuration.

        Expected payload:
        {
            "customer_name": "Customer Name",
            "project_name": "Project Name",
            "sites": [
                {
                    "name": "Site 1",
                    "address": "123 Main St",
                    "contact": {
                        "name": "Contact Name",
                        "email": "contact@email.com",
                        "phone": "123-456-7890",
                        "title": "Position"
                    },
                    "equipment": [
                        {
                            "type_id": 1,
                            "type_name": "SAN Switch",
                            "quantity": 2,
                            "items": [
                                {
                                    "switch_name": "SW01",
                                    "management_ip": "10.0.0.1",
                                    ...
                                }
                            ]
                        }
                    ]
                }
            ]
        }
        """
        try:
            data = request.data

            # Create Excel workbook
            wb = Workbook()
            # Remove the default sheet, we'll create one per site
            wb.remove(wb.active)

            # Define styles - Company Branding Colors
            # Color Palette:
            # #505050 - Dark Gray (headers, text)
            # #b3b3b3 - Light Gray (subheaders, borders)
            # #a2ca62 - Green (primary accent)
            # #64cae6 - Blue (secondary accent)
            # #ffffff - White (backgrounds, text on dark)

            header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='505050', end_color='505050', fill_type='solid')  # Dark Gray
            # Section headers: dark background with white text
            subheader_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')  # White text
            subheader_fill = PatternFill(start_color='505050', end_color='505050', fill_type='solid')  # Dark Gray background
            # Equipment detail headers: light gray (old subheader color)
            detail_header_font = Font(name='Calibri', size=11, bold=True, color='303030')  # Dark gray text
            detail_header_fill = PatternFill(start_color='B3B3B3', end_color='B3B3B3', fill_type='solid')  # Light Gray
            normal_font = Font(name='Calibri', size=11, color='505050')  # Dark Gray
            bold_font = Font(name='Calibri', size=11, bold=True, color='505050')  # Dark Gray
            light_gray_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')  # Light gray for field labels
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            thin_border = Border(
                left=Side(style='thin', color='B3B3B3'),
                right=Side(style='thin', color='B3B3B3'),
                top=Side(style='thin', color='B3B3B3'),
                bottom=Side(style='thin', color='B3B3B3')
            )

            # Loop through sites and create a sheet for each
            sites = data.get('sites', [])
            for site_index, site in enumerate(sites):
                # Create a new worksheet for this site
                ws = wb.create_sheet(title=site.get('name', f'Site {site_index + 1}')[:31])  # Sheet names max 31 chars

                # Hide gridlines for cleaner look
                ws.sheet_view.showGridLines = False

                current_row = 1

                # Pre-scan equipment to determine maximum column width for headers
                max_col = 4  # Default minimum width
                equipment_list = site.get('equipment', [])
                for equipment in equipment_list:
                    items = equipment.get('items', [])
                    if items and len(items) > 0:
                        fields = list(items[0].keys())
                        # Check if this is a switch (exclude VLAN for switches)
                        is_switch = 'switch' in equipment.get('type_name', '').lower()
                        # Subtract 1 from field count if it's a switch and has VLAN field
                        num_fields = len(fields)
                        if is_switch and 'vlan' in fields:
                            num_fields -= 1
                        max_col = max(max_col, num_fields)

                # Get the column letter for the rightmost column
                max_col_letter = get_column_letter(max_col)

                # Store logo path for later (will add to footer)
                logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'company-logo.png')

                # Start at row 1 (no blank rows at top)
                current_row = 1

                # Add title with dark background (first row) - size 26
                ws.merge_cells(f'A{current_row}:{max_col_letter}{current_row}')
                cell = ws[f'A{current_row}']
                cell.value = "Equipment Implementation Worksheet"
                cell.font = Font(name='Calibri', size=26, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='505050', end_color='505050', fill_type='solid')  # Dark gray from color scheme
                cell.alignment = center_alignment
                ws.row_dimensions[current_row].height = 40  # Proper height for size 26 font
                current_row += 2

                # Add project information section
                ws[f'A{current_row}'] = "Project Information"
                ws[f'A{current_row}'].font = subheader_font
                ws[f'A{current_row}'].fill = subheader_fill
                ws.merge_cells(f'A{current_row}:{max_col_letter}{current_row}')
                current_row += 1

                # Client Name - merge label across A:B, value starts at C
                ws[f'A{current_row}'] = "Client Name:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = data.get('customer_name', '')
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                current_row += 1

                # Project Name - merge label across A:B, value starts at C
                ws[f'A{current_row}'] = "Project:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = data.get('project_name', '')
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws[f'C{current_row}'].fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                current_row += 1

                # Planned Installation Date - merge label across A:B, value starts at C
                ws[f'A{current_row}'] = "Planned Installation Date:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = data.get('planned_installation_date', '')
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                current_row += 2

                # Site Information Section
                ws[f'A{current_row}'] = "Support Information"
                ws[f'A{current_row}'].font = subheader_font
                ws[f'A{current_row}'].fill = subheader_fill
                ws.merge_cells(f'A{current_row}:{max_col_letter}{current_row}')
                current_row += 1

                # Site Contact - Always include (empty or not) with borders and banding
                site_info_rows = []

                # Row 1 - Site Name - merge label across A:B, value starts at C
                ws[f'A{current_row}'] = "Site Name:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = site.get('name', '')
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                site_info_rows.append(current_row)
                current_row += 1

                # Row 2 - Site Contact - merge label across A:B, value starts at C
                ws[f'A{current_row}'] = "Site Contact:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = site.get('siteContactName', '')
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws[f'C{current_row}'].fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                site_info_rows.append(current_row)
                current_row += 1

                # Row 3 - Email - merge label across A:B, value starts at C
                ws[f'A{current_row}'] = "Email:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = site.get('siteContactEmail', '')
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                site_info_rows.append(current_row)
                current_row += 1

                # Row 4 - Phone - merge label across A:B, value starts at C
                phone_display = site.get('siteContactPhone', '')
                if site.get('siteContactAltPhone'):
                    phone_display += f" / {site.get('siteContactAltPhone', '')}"
                ws[f'A{current_row}'] = "Phone / Alt Phone:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = phone_display
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws[f'C{current_row}'].fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                site_info_rows.append(current_row)
                current_row += 1

                # Row 5 - Street Address - merge label across A:B, value starts at C
                ws[f'A{current_row}'] = "Street Address:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = site.get('siteStreetAddress', '')
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                site_info_rows.append(current_row)
                current_row += 1

                # Row 6 - City, State, Zip - merge label across A:B, value starts at C
                city_state_zip = site.get('siteCity', '')
                if site.get('siteState'):
                    city_state_zip += f", {site.get('siteState', '')}" if city_state_zip else site.get('siteState', '')
                if site.get('siteZip'):
                    city_state_zip += f" {site.get('siteZip', '')}"
                ws[f'A{current_row}'] = "City, State, Zip:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = city_state_zip
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws[f'C{current_row}'].fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                site_info_rows.append(current_row)
                current_row += 1

                # Row 7 - Location Notes - merge label across A:B, value starts at C
                ws[f'A{current_row}'] = "Location Notes:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = site.get('siteNotes', '')
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                site_info_rows.append(current_row)
                current_row += 1

                # Row 8 - DNS Servers - merge label across A:B, value starts at C
                dns_servers = site.get('dnsServer1', '')
                if site.get('dnsServer2'):
                    dns_servers += f", {site.get('dnsServer2', '')}" if dns_servers else site.get('dnsServer2', '')
                ws[f'A{current_row}'] = "DNS Servers:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = dns_servers
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws[f'C{current_row}'].fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                site_info_rows.append(current_row)
                current_row += 1

                # Row 9 - NTP Server - merge label across A:B, value starts at C
                ws[f'A{current_row}'] = "NTP Server:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = site.get('ntpServer', '')
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                site_info_rows.append(current_row)
                current_row += 1

                # Row 10 - SMTP Server - merge label across A:B, value starts at C
                smtp_display = site.get('smtpServer', '')
                if site.get('smtpPort'):
                    smtp_display += f":{site.get('smtpPort', '')}"
                ws[f'A{current_row}'] = "SMTP Server:Port:"
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].fill = light_gray_fill
                ws[f'A{current_row}'].border = thin_border
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'C{current_row}'] = smtp_display
                ws[f'C{current_row}'].font = normal_font
                ws[f'C{current_row}'].border = thin_border
                ws[f'C{current_row}'].fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                ws.merge_cells(f'C{current_row}:{max_col_letter}{current_row}')
                site_info_rows.append(current_row)
                current_row += 1

                current_row += 1

                # Add equipment sections for this site
                for equipment in equipment_list:
                    # Equipment type header - spans full width
                    ws[f'A{current_row}'] = f"{equipment.get('type_name', 'Equipment')} Details"
                    ws[f'A{current_row}'].font = subheader_font
                    ws[f'A{current_row}'].fill = subheader_fill
                    ws.merge_cells(f'A{current_row}:{max_col_letter}{current_row}')
                    current_row += 1

                    items = equipment.get('items', [])
                    if items and len(items) > 0:
                        equipment_type_name = equipment.get('type_name', 'Equipment').replace(' ', '_')
                        equipment_type_id = equipment.get('type_id')

                        # Check if this is a switch (exclude VLAN for switches)
                        is_switch = 'switch' in equipment.get('type_name', '').lower()

                        # Try to get field order from the equipment type's schema
                        # This ensures consistent ordering based on database configuration
                        try:
                            equipment_type = EquipmentType.objects.get(id=equipment_type_id)
                            schema_fields = [field['name'] for field in equipment_type.fields_schema]
                        except:
                            # Fallback to keys from first item if type not found
                            schema_fields = []

                        # Get all fields that actually exist in the items
                        all_fields = list(items[0].keys())

                        # Start with schema order if available, otherwise use all_fields
                        if schema_fields:
                            # Use schema order, but only include fields that actually exist in items
                            ordered_fields = [f for f in schema_fields if f in all_fields]
                            # Add any fields from items that aren't in schema (like network fields added by frontend)
                            for f in all_fields:
                                if f not in ordered_fields:
                                    ordered_fields.append(f)
                            all_fields = ordered_fields

                        # Reorder fields: put subnet_mask, default_gateway, vlan right after management_ip
                        # Exclude vlan for switches
                        if is_switch:
                            network_fields = ['subnet_mask', 'default_gateway']
                        else:
                            network_fields = ['subnet_mask', 'default_gateway', 'vlan']

                        fields = []

                        for field in all_fields:
                            # Skip vlan field for switches
                            if is_switch and field == 'vlan':
                                continue

                            if field == 'management_ip':
                                fields.append(field)
                                # Add network fields right after management_ip
                                for nf in network_fields:
                                    if nf in all_fields and nf not in fields:
                                        fields.append(nf)
                            elif field not in network_fields and field != 'vlan':
                                fields.append(field)

                        # Add any remaining network fields that weren't placed yet
                        for nf in network_fields:
                            if nf in all_fields and nf not in fields:
                                fields.append(nf)

                        # Special case label mappings for proper capitalization
                        label_overrides = {
                            'management_ip': 'Management IP',
                            'vlan': 'VLAN',
                            'subnet_mask': 'Subnet Mask',
                            'default_gateway': 'Default Gateway'
                        }

                        # Column labels row - Light gray (old subheader color)
                        for col_idx, field in enumerate(fields, start=1):
                            cell = ws.cell(row=current_row, column=col_idx)
                            # Use override label if available, otherwise convert snake_case to Title Case
                            if field in label_overrides:
                                label = label_overrides[field]
                            else:
                                label = field.replace('_', ' ').title()
                            cell.value = label
                            cell.font = detail_header_font  # Dark gray text
                            cell.fill = detail_header_fill  # Light gray background
                            cell.alignment = left_alignment  # Left-aligned to match data below
                            cell.border = thin_border
                        current_row += 1

                        # Data rows with alternating row colors and named ranges
                        for row_idx, item in enumerate(items):
                            item_number = row_idx + 1
                            for col_idx, field in enumerate(fields, start=1):
                                cell = ws.cell(row=current_row, column=col_idx)
                                cell.value = item.get(field, '')
                                cell.font = normal_font
                                cell.alignment = left_alignment
                                cell.border = thin_border
                                # Alternating row colors for better readability
                                if row_idx % 2 == 1:
                                    cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

                                # Create named range for this cell: equipment_type_fieldname_itemnumber
                                # e.g., "SAN_Switch_switch_name_1" or "Storage_Array_management_ip_2"
                                range_name = f"{equipment_type_name}_{field}_{item_number}"
                                # Excel named ranges have restrictions: max 255 chars, no spaces, must start with letter
                                range_name = range_name[:255]  # Truncate if too long
                                try:
                                    wb.defined_names[range_name] = f"'{ws.title}'!${get_column_letter(col_idx)}${current_row}"
                                except:
                                    pass  # Skip if name is invalid

                            current_row += 1

                        current_row += 1  # Blank row between equipment types

                # Add footer section
                current_row += 2  # Add some space

                # Blue-to-green gradient row before footer separator
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=current_row, column=col)
                    # Create gradient effect by transitioning from blue to green across columns
                    # Calculate color based on column position
                    progress = (col - 1) / max(max_col - 1, 1)  # 0.0 to 1.0
                    # Interpolate between blue (64CAE6) and green (A2CA62)
                    blue_r, blue_g, blue_b = 0x64, 0xCA, 0xE6
                    green_r, green_g, green_b = 0xA2, 0xCA, 0x62
                    r = int(blue_r + (green_r - blue_r) * progress)
                    g = int(blue_g + (green_g - blue_g) * progress)
                    b = int(blue_b + (green_b - blue_b) * progress)
                    color_hex = f"{r:02X}{g:02X}{b:02X}"
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                current_row += 1

                footer_start_row = current_row

                # Footer separator - full width of all columns
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = Border(top=Side(style='thick', color='505050'))
                current_row += 1

                # Implementation Team Contacts footer (centered text)
                implementation_team = site.get('implementationTeam', [])
                implementation_company_name = site.get('implementationCompanyName', '')
                contact_section_start_row = current_row  # Track where contact section starts
                contact_section_end_row = current_row  # Will be updated as we add rows

                if implementation_team and len(implementation_team) > 0:
                    ws.merge_cells(f'A{current_row}:{max_col_letter}{current_row}')
                    cell = ws[f'A{current_row}']
                    if implementation_company_name:
                        cell.value = f"Contact Us ({implementation_company_name})"
                    else:
                        cell.value = "Contact Us"
                    cell.font = Font(name='Calibri', size=10, bold=True, color='505050')
                    cell.alignment = center_alignment  # Center the text
                    current_row += 1

                    # Display each team member (centered)
                    for team_member in implementation_team:
                        ws.merge_cells(f'A{current_row}:{max_col_letter}{current_row}')
                        cell = ws[f'A{current_row}']
                        footer_text = f"{team_member.get('name', '')}"
                        if team_member.get('title'):
                            footer_text += f" - {team_member.get('title', '')}"
                        if team_member.get('email'):
                            footer_text += f" | {team_member.get('email', '')}"
                        if team_member.get('phone'):
                            footer_text += f" | {team_member.get('phone', '')}"
                        cell.value = footer_text
                        cell.font = Font(name='Calibri', size=9, color='505050')
                        cell.alignment = center_alignment
                        current_row += 1

                    # Update end of contact section (before timestamp)
                    contact_section_end_row = current_row - 1

                # Generated timestamp (centered) - separate from contact section
                ws.merge_cells(f'A{current_row}:{max_col_letter}{current_row}')
                cell = ws[f'A{current_row}']
                cell.value = f"Generated on {datetime.now().strftime('%Y-%m-%d at %H:%M')}"
                cell.font = Font(name='Calibri', size=8, italic=True, color='B3B3B3')
                cell.alignment = center_alignment

                # Add logo to right side of footer, vertically centered in contact section only
                if os.path.exists(logo_path) and implementation_team and len(implementation_team) > 0:
                    try:
                        img = XLImage(logo_path)
                        # Scale logo to fit nicely in footer
                        target_height = 60  # Slightly larger for better visibility
                        if img.height > 0:
                            aspect_ratio = img.width / img.height
                            img.height = target_height
                            img.width = int(target_height * aspect_ratio)

                        # Calculate vertical center position within contact section only
                        # For proper centering: if we have rows 1,2,3,4,5 the center is row 3 (index 2)
                        # Formula: start_row + (total_rows - 1) / 2
                        contact_section_height = contact_section_end_row - contact_section_start_row + 1
                        vertical_center_row = contact_section_start_row + ((contact_section_height - 1) // 2)

                        # Position logo on the right side (last column)
                        logo_col = max_col
                        logo_anchor = f'{get_column_letter(logo_col)}{vertical_center_row}'
                        img.anchor = logo_anchor
                        ws.add_image(img, logo_anchor)
                    except Exception as e:
                        print(f"Error loading logo in footer: {e}")

                # Auto-adjust column widths for this site's sheet
                # Track which columns contain network fields (IP, subnet, gateway) and name fields
                network_field_columns = set()
                name_field_columns = set()

                # First pass: identify special field columns
                for col_idx, column in enumerate(ws.columns, start=1):
                    column_letter = get_column_letter(col_idx)
                    for cell in column:
                        if cell.value and isinstance(cell.value, str):
                            cell_value_lower = cell.value.lower()
                            # Check if this is a network field header
                            if cell_value_lower in ['management ip', 'subnet mask', 'default gateway', 'ip address', 'gateway']:
                                network_field_columns.add(column_letter)
                                break
                            # Check if this is a name field header (switch name, array name, etc.)
                            elif 'name' in cell_value_lower and cell_value_lower not in ['site name:', 'client name:', 'site contact:']:
                                name_field_columns.add(column_letter)
                                break

                # Second pass: set column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                    # Calculate base width
                    adjusted_width = max_length + 2

                    # Apply 25% increase for network field columns
                    if column_letter in network_field_columns:
                        adjusted_width = int(adjusted_width * 1.25)
                        # For empty columns, ensure minimum width for IP addresses
                        if max_length < 10:  # Empty or very short content
                            adjusted_width = 20  # Wide enough for "255.255.255.255"

                    # Apply 50% increase for name field columns (e.g., Switch Name, Array Name)
                    elif column_letter in name_field_columns:
                        adjusted_width = int(adjusted_width * 1.5)
                        # For empty columns, ensure minimum width for device names
                        if max_length < 15:  # Empty or very short content
                            adjusted_width = 25  # Wide enough for longer device names

                    # Cap at 50 characters
                    adjusted_width = min(adjusted_width, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Create response
            filename = f"implementation_worksheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            return Response(
                {"error": f"Failed to generate worksheet: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
